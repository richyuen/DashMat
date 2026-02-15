from __future__ import annotations

import base64
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd
import pytest
import utils.parsing as parsing

from utils.parsing import (
    convert_percents_to_decimals,
    detect_periodicity,
    get_sheet_names,
    parse_uploaded_file,
    parse_uploaded_sheets,
)


def _as_upload_payload(text: str) -> str:
    encoded = base64.b64encode(text.encode("utf-8")).decode("ascii")
    return f"data:text/csv;base64,{encoded}"


def _as_upload_excel(df: pd.DataFrame) -> str:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
    encoded = base64.b64encode(bio.getvalue()).decode("ascii")
    return f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{encoded}"


def _as_upload_openpyxl_workbook(workbook: Workbook) -> str:
    bio = BytesIO()
    workbook.save(bio)
    encoded = base64.b64encode(bio.getvalue()).decode("ascii")
    return f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{encoded}"


def _decode_payload(payload: str) -> bytes:
    return base64.b64decode(payload.split(",", 1)[1])


def _build_morningstar_workbook(*, daily: bool) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Morningstar Export"
    ws["A2"] = "Currency: Base Currency"
    ws["A3"] = "Grouped by: Morningstar Category"
    ws["A4"] = "Calculated on: 2/15/2026 10:00:00 AM"
    ws["A5"] = "Exported on: 2/15/2026 10:00:01 AM"

    if daily:
        start_1 = end_1 = "1/1/2024"
        start_2 = end_2 = "1/2/2024"
    else:
        start_1, end_1 = "1/1/2024", "1/31/2024"
        start_2, end_2 = "2/1/2024", "2/29/2024"

    ws.merge_cells("C7:D7")
    ws["C7"] = start_1
    ws.merge_cells("E7:F7")
    ws["E7"] = start_2
    ws["G7"] = start_2

    ws.merge_cells("C8:D8")
    ws["C8"] = end_1
    ws.merge_cells("E8:F8")
    ws["E8"] = end_2
    ws["G8"] = end_2

    ws["A9"] = "Group/Investment"
    ws["B9"] = "SecId"
    ws["C9"] = "Return"
    ws["D9"] = "Peer group percentile"
    ws["E9"] = "Return"
    ws["F9"] = "Peer group percentile"
    ws["G9"] = "Longest Up-Streak Return"
    ws["H9"] = "Return Date (Daily)"

    ws["A11"] = "US Fund Target-Date 2030"
    ws["A11"].font = Font(bold=True)

    ws["A12"] = "Fund Alpha"
    ws["B12"] = "SEC001"
    ws["C12"] = 1.5
    ws["D12"] = 1
    ws["E12"] = -0.5
    ws["F12"] = 2
    ws["G12"] = 99.9
    ws["H12"] = "2024-01-02"

    ws["A13"] = "Display Group 25th Percentile"
    ws["A13"].font = Font(italic=True)
    ws["C13"] = 0.25
    ws["E13"] = 0.5

    ws["A14"] = "Fund Beta"
    ws["B14"] = "SEC002"
    ws["C14"] = "2%"
    ws["D14"] = 3
    ws["E14"] = "1%"
    ws["F14"] = 4
    ws["G14"] = 88.8
    ws["H14"] = "2024-01-02"

    return wb


def test_convert_percents_to_decimals_handles_percent_strings():
    df = pd.DataFrame({"A": ["5%", "2.5%", "-1%"], "B": [0.01, 0.02, 0.03]})
    converted = convert_percents_to_decimals(df)
    assert converted["A"].iloc[0] == pytest.approx(0.05)
    assert converted["A"].iloc[1] == pytest.approx(0.025)
    assert converted["A"].iloc[2] == pytest.approx(-0.01)
    assert converted["B"].iloc[1] == pytest.approx(0.02)


def test_detect_periodicity_distinguishes_daily_and_monthly():
    daily_idx = pd.date_range("2024-01-01", periods=5, freq="B")
    monthly_idx = pd.date_range("2024-01-31", periods=5, freq="ME")
    daily_df = pd.DataFrame({"x": range(5)}, index=daily_idx)
    monthly_df = pd.DataFrame({"x": range(5)}, index=monthly_idx)

    assert detect_periodicity(daily_df) == "daily"
    assert detect_periodicity(monthly_df) == "monthly"


def test_detect_periodicity_defaults_to_daily_for_single_row():
    df = pd.DataFrame({"x": [1]}, index=pd.to_datetime(["2024-01-01"]))
    assert detect_periodicity(df) == "daily"


def test_parse_uploaded_file_csv_sorts_dates_and_converts_percents():
    csv = "Date,A,B\n2024-01-03,1%,0.01\n2024-01-01,2%,0.02\n"
    payload = _as_upload_payload(csv)

    parsed = parse_uploaded_file(payload, "returns.csv")

    assert list(parsed.columns) == ["A", "B"]
    assert parsed.index[0].strftime("%Y-%m-%d") == "2024-01-01"
    assert parsed["A"].iloc[0] == pytest.approx(0.02)
    assert parsed["A"].iloc[1] == pytest.approx(0.01)


def test_parse_uploaded_sheets_csv_respects_requested_key_name():
    csv = "Date,A\n2024-01-01,1%\n"
    payload = _as_upload_payload(csv)

    parsed = parse_uploaded_sheets(payload, "returns.csv", ["Custom"])

    assert list(parsed.keys()) == ["Custom"]
    assert parsed["Custom"].iloc[0, 0] == pytest.approx(0.01)


def test_parse_uploaded_sheets_csv_rejects_multi_sheet_requests():
    csv = "Date,A\n2024-01-01,1%\n"
    payload = _as_upload_payload(csv)
    with pytest.raises(ValueError, match="do not support multiple sheets"):
        parse_uploaded_sheets(payload, "returns.csv", ["S1", "S2"])


def test_parse_uploaded_file_raises_for_unsupported_extension():
    payload = _as_upload_payload("Date,A\n2024-01-01,0.01\n")
    with pytest.raises(ValueError):
        parse_uploaded_file(payload, "returns.txt")


def test_get_sheet_names_returns_empty_for_non_excel():
    payload = _as_upload_payload("Date,A\n2024-01-01,0.01\n")
    assert get_sheet_names(payload, "returns.csv") == []


def test_get_sheet_names_and_parse_uploaded_file_excel():
    df = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2024-01-02", "2024-01-01"]),
            "A": ["1%", "2%"],
        }
    )
    payload = _as_upload_excel(df)

    sheets = get_sheet_names(payload, "returns.xlsx")
    parsed = parse_uploaded_file(payload, "returns.xlsx", sheet_name=0)

    assert sheets == ["Sheet1"]
    assert parsed.index[0].strftime("%Y-%m-%d") == "2024-01-01"
    assert parsed["A"].iloc[0] == pytest.approx(0.02)


def test_parse_uploaded_sheets_excel_resolves_indices_and_deduplicates_requests():
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        pd.DataFrame({"Date": ["2024-01-01"], "A": [0.1]}).to_excel(writer, sheet_name="S1", index=False)
        pd.DataFrame({"Date": ["2024-01-02"], "B": [0.2]}).to_excel(writer, sheet_name="S2", index=False)
    payload = (
        "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
        + base64.b64encode(bio.getvalue()).decode("ascii")
    )

    parsed = parse_uploaded_sheets(payload, "multi.xlsx", [0, "S2", 0])

    assert list(parsed.keys()) == ["S1", "S2"]
    assert list(parsed["S1"].columns) == ["A"]
    assert list(parsed["S2"].columns) == ["B"]


def test_parse_uploaded_sheets_ignore_errors_skips_bad_sheet():
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        pd.DataFrame({"Date": ["2024-01-01"], "A": [0.1]}).to_excel(writer, sheet_name="Good", index=False)
        pd.DataFrame({"Date": ["not-a-date"], "A": [0.2]}).to_excel(writer, sheet_name="Bad", index=False)
    payload = (
        "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
        + base64.b64encode(bio.getvalue()).decode("ascii")
    )

    parsed = parse_uploaded_sheets(payload, "mixed.xlsx", ["Good", "Bad"], ignore_errors=True)

    assert list(parsed.keys()) == ["Good"]
    assert parsed["Good"].shape == (1, 1)


def test_parse_uploaded_sheets_ignore_errors_raises_when_all_bad():
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        pd.DataFrame({"Date": ["not-a-date"], "A": [0.2]}).to_excel(writer, sheet_name="Bad", index=False)
    payload = (
        "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
        + base64.b64encode(bio.getvalue()).decode("ascii")
    )

    with pytest.raises(Exception):
        parse_uploaded_sheets(payload, "bad.xlsx", ["Bad"], ignore_errors=True)


def test_parse_uploaded_file_morningstar_daily_imports_only_exact_return_columns():
    payload = _as_upload_openpyxl_workbook(_build_morningstar_workbook(daily=True))

    parsed = parse_uploaded_file(payload, "morningstar.xlsx")

    assert list(parsed.columns) == ["Fund Alpha", "Fund Beta"]
    assert list(parsed.index) == list(pd.to_datetime(["2024-01-01", "2024-01-02"]))
    assert parsed.loc[pd.Timestamp("2024-01-01"), "Fund Alpha"] == pytest.approx(0.015)
    assert parsed.loc[pd.Timestamp("2024-01-02"), "Fund Alpha"] == pytest.approx(-0.005)
    assert parsed.loc[pd.Timestamp("2024-01-01"), "Fund Beta"] == pytest.approx(0.02)
    assert parsed.loc[pd.Timestamp("2024-01-02"), "Fund Beta"] == pytest.approx(0.01)
    assert "US Fund Target-Date 2030" not in parsed.columns
    assert "Display Group 25th Percentile" not in parsed.columns
    assert float(parsed.max().max()) < 0.1
    assert detect_periodicity(parsed) == "daily"


def test_parse_uploaded_file_morningstar_monthly_uses_first_date_pair_for_periodicity():
    payload = _as_upload_openpyxl_workbook(_build_morningstar_workbook(daily=False))

    parsed = parse_uploaded_file(payload, "morningstar_monthly.xlsx")

    assert list(parsed.index) == list(pd.to_datetime(["2024-01-31", "2024-02-29"]))
    assert detect_periodicity(parsed) == "monthly"


def test_parse_uploaded_file_raises_for_empty_standard_table():
    payload = _as_upload_payload("Date,A\n")
    with pytest.raises(ValueError, match="contains no data"):
        parse_uploaded_file(payload, "empty.csv")


def test_parse_uploaded_file_falls_back_to_standard_excel_when_not_morningstar():
    df = pd.DataFrame(
        {
            "Date": pd.to_datetime(["2024-01-03", "2024-01-01"]),
            "A": [0.02, 0.01],
        }
    )
    payload = _as_upload_excel(df)
    parsed = parse_uploaded_file(payload, "plain.xlsx")
    assert list(parsed.index) == list(pd.to_datetime(["2024-01-01", "2024-01-03"]))
    assert list(parsed.columns) == ["A"]


def test_internal_helpers_cover_blank_invalid_and_numeric_branches():
    assert parsing._is_blank(None) is True
    assert parsing._is_blank("   ") is True
    assert parsing._is_blank(0) is False

    assert parsing._to_header_text(None) == ""
    assert parsing._to_header_text(" x ") == "x"

    assert parsing._to_timestamp("") is None
    assert parsing._to_timestamp("not-a-date") is None
    assert parsing._to_timestamp("2024-01-01") == pd.Timestamp("2024-01-01")

    assert pd.isna(parsing._coerce_morningstar_return(None))
    assert parsing._coerce_morningstar_return("2%") == pytest.approx(0.02)
    assert parsing._coerce_morningstar_return("1,234.5") == pytest.approx(12.345)
    assert pd.isna(parsing._coerce_morningstar_return("not-a-number"))


def test_resolve_sheet_name_accepts_string():
    wb = Workbook()
    wb.active.title = "SheetX"
    assert parsing._resolve_sheet_name(wb, "SheetX") == "SheetX"
    assert parsing._resolve_sheet_name(wb, 0) == "SheetX"


def test_resolve_requested_sheet_names_defaults_and_rejects_missing():
    wb = Workbook()
    wb.active.title = "First"
    wb.create_sheet("Second")

    assert parsing._resolve_requested_sheet_names(wb, []) == ["First"]
    assert parsing._resolve_requested_sheet_names(wb, [0, "Second", 0]) == ["First", "Second"]
    with pytest.raises(ValueError, match="Sheet not found"):
        parsing._resolve_requested_sheet_names(wb, ["Missing"])


def test_find_morningstar_header_row_returns_none_when_missing():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "No matching header"
    assert parsing._find_morningstar_header_row(ws) is None


def test_convert_percents_to_decimals_handles_object_without_percent():
    df = pd.DataFrame({"A": ["1.2", "3.4"], "B": [1, 2]})
    converted = convert_percents_to_decimals(df)
    assert converted["A"].iloc[0] == pytest.approx(1.2)
    assert converted["B"].iloc[1] == pytest.approx(2.0)


def test_parse_morningstar_raises_for_header_too_early():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A2"] = "Group/Investment"
    ws["B2"] = "Return"
    ws["A3"] = "Fund"
    ws["B3"] = 1.0

    payload = _as_upload_openpyxl_workbook(wb)
    with pytest.raises(ValueError, match="Unexpected Morningstar header layout"):
        parsing._parse_morningstar_report(_decode_payload(payload), "Sheet1")


def test_parse_morningstar_raises_for_no_dated_return_columns():
    wb = _build_morningstar_workbook(daily=True)
    ws = wb.active
    # Remove both date rows, leaving Return columns with no associated dates.
    ws["C7"] = None
    ws["E7"] = None
    ws["G7"] = None
    ws["C8"] = None
    ws["E8"] = None
    ws["G8"] = None

    payload = _as_upload_openpyxl_workbook(wb)
    with pytest.raises(ValueError, match="No dated Return columns"):
        parsing._parse_morningstar_report(_decode_payload(payload), "Sheet1")


def test_parse_morningstar_raises_for_no_importable_rows():
    wb = _build_morningstar_workbook(daily=True)
    ws = wb.active
    # Remove all numeric data rows.
    for row in (12, 14):
        ws[f"C{row}"] = None
        ws[f"E{row}"] = None

    payload = _as_upload_openpyxl_workbook(wb)
    with pytest.raises(ValueError, match="No importable Morningstar data rows"):
        parsing._parse_morningstar_report(_decode_payload(payload), "Sheet1")


def test_parse_morningstar_handles_duplicate_dates_and_latest_duplicate_row_wins():
    wb = _build_morningstar_workbook(daily=True)
    ws = wb.active
    # Force duplicate period dates across both Return columns.
    ws["E8"] = ws["C8"].value
    # Add duplicate series row with later non-null value for first date.
    ws["A15"] = "Fund Alpha"
    ws["C15"] = 9.9
    ws["E15"] = None

    parsed = parsing._parse_morningstar_report(_decode_payload(_as_upload_openpyxl_workbook(wb)), "Sheet1")

    assert len(parsed.index.unique()) == len(parsed.index)
    assert parsed.loc[pd.Timestamp("2024-01-01"), "Fund Alpha"] == pytest.approx(-0.005)


def test_parse_morningstar_keeps_periodicity_hint_none_when_first_pair_partial():
    wb = _build_morningstar_workbook(daily=True)
    ws = wb.active
    # Keep end dates only so hint cannot be determined.
    ws["C7"] = None
    ws["E7"] = None

    parsed = parsing._parse_morningstar_report(_decode_payload(_as_upload_openpyxl_workbook(wb)), "Sheet1")
    assert parsed.attrs.get("periodicity_hint") is None


def test_detect_periodicity_uses_hint_over_index_spacing():
    idx = pd.date_range("2024-01-31", periods=3, freq="ME")
    df = pd.DataFrame({"A": [0.01, 0.02, 0.03]}, index=idx)
    df.attrs["periodicity_hint"] = "daily"
    assert detect_periodicity(df) == "daily"


def test_parse_morningstar_missing_group_and_return_branches(monkeypatch):
    payload = _as_upload_openpyxl_workbook(_build_morningstar_workbook(daily=True))
    decoded = _decode_payload(payload)

    # Force a header row without Group/Investment.
    def _header_row_with_no_group(_ws):
        return 9

    monkeypatch.setattr(parsing, "_find_morningstar_header_row", _header_row_with_no_group)

    original_to_header_text = parsing._to_header_text

    def _no_group_header(value):
        text = original_to_header_text(value)
        return "NotGroup" if text == "Group/Investment" else text

    monkeypatch.setattr(parsing, "_to_header_text", _no_group_header)
    with pytest.raises(ValueError, match="Missing Group/Investment"):
        parsing._parse_morningstar_report(decoded, "Sheet1")


def test_parse_morningstar_no_exact_return_columns_branch(monkeypatch):
    payload = _as_upload_openpyxl_workbook(_build_morningstar_workbook(daily=True))
    decoded = _decode_payload(payload)

    monkeypatch.setattr(parsing, "_find_morningstar_header_row", lambda _ws: 9)
    original_to_header_text = parsing._to_header_text

    def _rename_return(value):
        text = original_to_header_text(value)
        return "ReturnX" if text == "Return" else text

    monkeypatch.setattr(parsing, "_to_header_text", _rename_return)
    with pytest.raises(ValueError, match="No exact Return columns"):
        parsing._parse_morningstar_report(decoded, "Sheet1")


def test_parse_morningstar_stripped_series_name_empty_branch(monkeypatch):
    wb = _build_morningstar_workbook(daily=True)
    ws = wb.active
    ws["A12"] = "   "
    ws["A14"] = "   "

    monkeypatch.setattr(parsing, "_is_blank", lambda value: value is None)
    with pytest.raises(ValueError, match="No importable Morningstar data rows"):
        parsing._parse_morningstar_report(
            _decode_payload(_as_upload_openpyxl_workbook(wb)),
            "Sheet1",
        )
