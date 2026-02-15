"""Parsing utilities for CSV/Excel files with returns data."""

import base64
import io
from typing import Any

from openpyxl import load_workbook
import pandas as pd


class _MorningstarLayoutNotFoundError(ValueError):
    """Raised when a workbook sheet is not in Morningstar report layout."""


def get_sheet_names(contents: str, filename: str) -> list[str]:
    """Return the list of sheet names for an Excel file.

    Args:
        contents: Base64 encoded file contents from dcc.Upload
        filename: Original filename to determine file type

    Returns:
        List of sheet names, or empty list for non-Excel files.
    """
    if not filename.endswith((".xlsx", ".xls")):
        return []
    _content_type, content_string = contents.split(",")
    decoded = base64.b64decode(content_string)
    xls = pd.ExcelFile(io.BytesIO(decoded))
    return xls.sheet_names


def parse_uploaded_file(contents: str, filename: str, sheet_name=0) -> pd.DataFrame:
    """Parse uploaded file contents into a DataFrame.

    Args:
        contents: Base64 encoded file contents from dcc.Upload
        filename: Original filename to determine file type
        sheet_name: Sheet name or index for Excel files (default: 0, first sheet)

    Returns:
        DataFrame with DatetimeIndex and returns as columns
    """
    parsed = parse_uploaded_sheets(contents, filename, [sheet_name])
    return next(iter(parsed.values()))


def parse_uploaded_sheets(
    contents: str,
    filename: str,
    sheet_names: list[int | str],
    ignore_errors: bool = False,
) -> dict[str, pd.DataFrame]:
    """Parse one or more uploaded sheets with a single decode/load pass.

    Args:
        contents: Base64 encoded file contents from dcc.Upload
        filename: Original filename to determine file type
        sheet_names: Requested sheet names/indices
        ignore_errors: If True, skip per-sheet parse errors and continue.

    Returns:
        Mapping of resolved sheet name -> parsed returns DataFrame.
    """
    _content_type, content_string = contents.split(",")
    decoded = base64.b64decode(content_string)

    if filename.endswith(".csv"):
        if len(sheet_names) > 1:
            raise ValueError("CSV uploads do not support multiple sheets.")
        df = pd.read_csv(io.StringIO(decoded.decode("utf-8")))
        key = str(sheet_names[0]) if sheet_names else "Sheet1"
        return {key: _normalize_standard_returns_df(df)}

    if not filename.endswith((".xlsx", ".xls")):
        raise ValueError(f"Unsupported file type: {filename}")

    wb = load_workbook(io.BytesIO(decoded), data_only=True)
    xls = pd.ExcelFile(io.BytesIO(decoded))
    requested = _resolve_requested_sheet_names(wb, sheet_names)

    results: dict[str, pd.DataFrame] = {}
    first_error = None
    for resolved_sheet in requested:
        try:
            results[resolved_sheet] = _parse_excel_sheet_from_context(wb, xls, resolved_sheet)
        except Exception as exc:
            if first_error is None:
                first_error = exc
            if not ignore_errors:
                raise

    if not results:
        if first_error is not None:
            raise first_error
        raise ValueError("No sheets were parsed.")

    return results


def _resolve_requested_sheet_names(workbook, sheet_names: list[int | str]) -> list[str]:
    if not sheet_names:
        return [workbook.sheetnames[0]]

    resolved: list[str] = []
    seen = set()
    for raw in sheet_names:
        name = _resolve_sheet_name(workbook, raw)
        if name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {name}")
        if name not in seen:
            resolved.append(name)
            seen.add(name)
    return resolved


def _parse_excel_sheet_from_context(workbook, excel_file: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    worksheet = workbook[sheet_name]
    try:
        return _parse_morningstar_report_from_worksheet(worksheet)
    except _MorningstarLayoutNotFoundError:
        standard_df = excel_file.parse(sheet_name=sheet_name)
        return _normalize_standard_returns_df(standard_df)


def _normalize_standard_returns_df(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize a standard date-indexed returns table."""
    if df.empty:
        raise ValueError("Uploaded file contains no data.")

    date_col = df.columns[0]
    result = df.copy()
    result[date_col] = pd.to_datetime(result[date_col], errors="raise")
    result = result.set_index(date_col)
    result.index.name = "Date"
    result = convert_percents_to_decimals(result)
    result = result.sort_index()
    return result


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _to_header_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_timestamp(value: Any) -> pd.Timestamp | None:
    if _is_blank(value):
        return None
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return None
    return pd.Timestamp(ts)


def _forward_fill_row_values(ws, row_idx: int, max_col: int) -> list[Any]:
    values: list[Any] = []
    last_nonblank = None
    for col_idx in range(1, max_col + 1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if _is_blank(value):
            values.append(last_nonblank)
        else:
            last_nonblank = value
            values.append(value)
    return values


def _coerce_morningstar_return(value: Any) -> float:
    """Convert Morningstar return cells to decimal returns.

    Morningstar numeric values in this report format are percent units.
    """
    if _is_blank(value):
        return float("nan")

    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text.endswith("%"):
            text = text[:-1].strip()
        number = pd.to_numeric(text, errors="coerce")
    else:
        number = pd.to_numeric(value, errors="coerce")

    if pd.isna(number):
        return float("nan")

    return float(number) / 100.0


def _resolve_sheet_name(workbook, sheet_name: int | str) -> str:
    if isinstance(sheet_name, int):
        return workbook.sheetnames[sheet_name]
    return sheet_name


def _find_morningstar_header_row(ws) -> int | None:
    max_scan_row = min(ws.max_row, 250)
    for row_idx in range(1, max_scan_row + 1):
        row_values = [_to_header_text(ws.cell(row=row_idx, column=c).value) for c in range(1, ws.max_column + 1)]
        if "Group/Investment" in row_values and any(v == "Return" for v in row_values):
            return row_idx
    return None


def _parse_morningstar_report(decoded: bytes, sheet_name: int | str) -> pd.DataFrame:
    wb = load_workbook(io.BytesIO(decoded), data_only=True)
    ws = wb[_resolve_sheet_name(wb, sheet_name)]
    return _parse_morningstar_report_from_worksheet(ws)


def _parse_morningstar_report_from_worksheet(ws) -> pd.DataFrame:
    header_row = _find_morningstar_header_row(ws)
    if header_row is None:
        raise _MorningstarLayoutNotFoundError("Unsupported Excel layout.")
    if header_row < 3:
        raise ValueError("Unexpected Morningstar header layout.")

    start_row = header_row - 2
    end_row = header_row - 1
    header_values = [_to_header_text(ws.cell(row=header_row, column=c).value) for c in range(1, ws.max_column + 1)]

    try:
        group_col = header_values.index("Group/Investment") + 1
    except ValueError as exc:
        raise ValueError("Missing Group/Investment column in Morningstar report.") from exc

    return_cols = [idx + 1 for idx, value in enumerate(header_values) if value == "Return"]
    if not return_cols:
        raise ValueError("No exact Return columns found in Morningstar report.")

    start_values = _forward_fill_row_values(ws, start_row, ws.max_column)
    end_values = _forward_fill_row_values(ws, end_row, ws.max_column)

    dated_return_cols: list[int] = []
    period_dates: list[pd.Timestamp] = []
    periodicity_hint = None

    for col_idx in return_cols:
        start_ts = _to_timestamp(start_values[col_idx - 1])
        end_ts = _to_timestamp(end_values[col_idx - 1])
        period_ts = end_ts or start_ts
        if period_ts is None:
            continue
        if periodicity_hint is None and start_ts is not None and end_ts is not None:
            periodicity_hint = "daily" if start_ts.normalize() == end_ts.normalize() else "monthly"
        dated_return_cols.append(col_idx)
        period_dates.append(period_ts)

    if not dated_return_cols:
        raise ValueError("No dated Return columns found in Morningstar report.")

    period_index = pd.DatetimeIndex(period_dates)
    series_map: dict[str, pd.Series] = {}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=group_col)
        raw_name = name_cell.value
        if _is_blank(raw_name):
            continue

        series_name = str(raw_name).strip()
        if not series_name:
            continue

        if bool(name_cell.font.bold) or bool(name_cell.font.italic):
            continue

        row_values = [
            _coerce_morningstar_return(ws.cell(row=row_idx, column=col_idx).value)
            for col_idx in dated_return_cols
        ]
        row_series = pd.Series(row_values, index=period_index, dtype="float64")
        if row_series.notna().sum() == 0:
            continue

        existing = series_map.get(series_name)
        if existing is None:
            series_map[series_name] = row_series
        else:
            # Prefer later rows when they provide non-null values.
            series_map[series_name] = row_series.combine_first(existing)

    if not series_map:
        raise ValueError("No importable Morningstar data rows found.")

    result = pd.DataFrame(series_map)
    if result.index.has_duplicates:
        result = result.groupby(level=0).last()
    result.index.name = "Date"
    result = result.sort_index()

    if periodicity_hint in {"daily", "monthly"}:
        result.attrs["periodicity_hint"] = periodicity_hint

    return result


def convert_percents_to_decimals(df: pd.DataFrame) -> pd.DataFrame:
    """Convert any percent-formatted values to decimals.

    Detects values with '%' suffix and divides by 100.
    """
    result = df.copy()

    for col in result.columns:
        if pd.api.types.is_object_dtype(result[col]) or pd.api.types.is_string_dtype(result[col]):
            str_series = result[col].astype(str)
            mask = str_series.str.contains("%", na=False)
            if mask.any():
                result[col] = (
                    str_series
                    .str.replace("%", "", regex=False)
                    .astype(float)
                    / 100
                )

    for col in result.columns:
        result[col] = pd.to_numeric(result[col], errors="coerce")

    return result


def detect_periodicity(df: pd.DataFrame) -> str:
    """Detect if the data is daily or monthly using the first few rows.

    Returns:
        'daily' or 'monthly'
    """
    hint = df.attrs.get("periodicity_hint")
    if hint in {"daily", "monthly"}:
        return hint

    if len(df) < 2:
        return "daily"

    sample_index = df.index[:5]
    date_diffs = pd.Series(sample_index).diff().dropna()
    median_diff = date_diffs.median().days

    if median_diff > 20:
        return "monthly"
    return "daily"
